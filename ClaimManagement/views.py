import calendar
from openpyxl import load_workbook 
from openpyxl.styles import Font
import xlrd
from openpyxl.utils import get_column_letter
import openpyxl
from rest_framework import generics
from .serializers import *
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from .models import *
from django.core.files.uploadedfile import SimpleUploadedFile
import os
import datetime
import zipfile
import pandas as pd
import io
from rest_framework import status
from PreAuth.views import PDFGenerator
from rest_framework.views import APIView
from collections import OrderedDict
from rest_framework.pagination import LimitOffsetPagination
import csv
from io import StringIO
from PreAuth.views import logger
from django.core.files import File
from rest_framework.filters import SearchFilter
# Create your views here.

# This is for .CSV files
class ExcelMergeCSVAPIView(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = ExcelFileSerializer

    def post(self, request):
        serializer = ExcelFileSerializer(data=request.data)
        if serializer.is_valid():
            # Load the CSV files
            csv_files = request.FILES.getlist('file')
            if len(csv_files) < 2:
                return Response({'message': 'At least two CSV files are required for Merging.',
                                 'status': 'error'},
                                status=status.HTTP_400_BAD_REQUEST)

            # Merge the CSV files
            dfs = [pd.read_csv(f, header=2) for f in csv_files]
            merged_df = pd.concat(dfs, ignore_index=True)

            # Write the merged data to an XLSX file
            output_file = io.BytesIO()
            writer = pd.ExcelWriter(output_file, engine='openpyxl')
            merged_df.to_excel(writer, index=False, sheet_name='Sheet1')

            # Add the header only once
            worksheet = writer.sheets['Sheet1']
            for i, col in enumerate(merged_df.columns):
                worksheet.cell(row=1, column=i+1, value=col)

            # Save the XLSX file
            writer.save()
            output_file.seek(0)
            date = datetime.datetime.today().date()
            # Return the merged file as a response
            response = HttpResponse(output_file.read(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename= MERGE_DUMP_{date}.xlsx'.format(
                date=date)
            return response

# This is for .xls files
class ExcelMergexlsAPIView(generics.GenericAPIView):
    serializer_class = ExcelFileSerializer
    parser_classes = [MultiPartParser]

    def post(self, request, format=None):
        serializer = ExcelFileSerializer(data=request.data)
        if serializer.is_valid():
            # Load the Excel files
            excel_files = []
            for file in request.FILES.getlist('file'):
                excel_files.append(xlrd.open_workbook(
                    file_contents=file.read()))
            if len(excel_files) == 1:
                return Response({'status': 'error',
                                'message': "To Merge the Excel's two or more .xls files will be needed."})
            # Merge the Excel files

            merged_workbook = openpyxl.Workbook()
            merged_sheet = merged_workbook.active
            merged_sheet.title = 'Merged Sheet'
            header_added = False  # To keep track of whether the header has been added
            row_index_merged = 1  # To keep track of the row index in the merged sheet
            for excel_file in excel_files:
                for sheet in excel_file.sheets():
                    for row_index in range(sheet.nrows):
                        # If the sheet is the first sheet, and header has not been added yet
                        if not header_added and row_index == 2:
                            # Add the headers to the merged sheet
                            for col_index in range(sheet.ncols):
                                column_letter = get_column_letter(col_index+1)
                                merged_sheet[column_letter+str(row_index_merged)] = sheet.cell_value(
                                    row_index, col_index)
                                merged_sheet[column_letter +
                                             str(row_index_merged)].font = Font(bold=True)
                            row_index_merged += 1
                            header_added = True  # Set header_added to True so that it is not added again
                        elif row_index > 2:  # If the row is after the header row
                            for col_index in range(sheet.ncols):
                                column_letter = get_column_letter(col_index+1)
                                merged_sheet[column_letter+str(row_index_merged)] = sheet.cell_value(
                                    row_index, col_index)
                            row_index_merged += 1

            # Write the merged Excel file to a BytesIO object
            response = HttpResponse(
                content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=MERGE_DUMP.xlsx'
            merged_workbook.save(response)

            return response
        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'})


class CaseSheetPostForm(generics.GenericAPIView):
    serializer_class = CaseSheetSerializer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        caseNumber = request.data['caseNumber'].replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['admitCaseClinicalSheet', 'icp', 'medicationChart', 'initialSheet', 'tprChart',
                      'vitalChart', 'losStayChart', 'caseSheetOtherDocuments']:

            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 2000 * 1024:
                        return Response({'status': 'error', 'message': 'Each Image size should be less than 1.99 MB'}, status=400)
                pdf = PDFGenerator(images, caseNumber, field)
                files[field] = SimpleUploadedFile(
                    f"{caseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

        serializer = self.get_serializer(data={
            'caseNumber': request.data['caseNumber'],
            **files})

        if serializer.is_valid():
            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data saved successfully'}, status=200)
        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)


class CaseSheetUpdateForm(generics.GenericAPIView):
    serializer_class = CaseSheetUpdateSerializer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    # def put(self , request , caseNumber ):
    #     try:
    #         instance = CaseSheet.objects.get(caseNumber__caseNumber = caseNumber)
    
    #     except:
    #         return Response({'status': 'error',
    #                          'message' : 'caseNumber ID is not found'}, status=400)

    #     caseNumber =caseNumber.replace('/', '')
    #     date = datetime.datetime.today().date()
    #     files = {}
    #     for field in ['admitCaseClinicalSheet', 'icp', 'medicationChart', 'initialSheet', 'tprChart',
    #                   'vitalChart', 'losStayChart', 'caseSheetOtherDocuments']:

    #         images = request.FILES.getlist(field)
    #         if images:
    #             for image in images:
    #                     if image.size > 2000 * 1024:
    #                         return Response({'status': 'error',
    #                                          'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
    #                                          status=400 )
    #             pdf = PDFGenerator(images, caseNumber, field)
    #             files[field] = SimpleUploadedFile(
    #                 f"{caseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

    #     serializer = self.get_serializer(instance,data={
    #         'admitCaseClinicalSheet': files.get('admitCaseClinicalSheet', None),
    #         'icp': files.get('icp', None),
    #         'medicationChart': files.get('medicationChart', None),
    #         'initialSheet': files.get('initialSheet', None),
    #         'tprChart': files.get('tprChart', None),
    #         'vitalChart': files.get('vitalChart', None),
    #         'losStayChart': files.get('losStayChart', None),
    #         'caseSheetOtherDocuments': files.get('caseSheetOtherDocuments', None),
    #          } , partial = False)

    #     if serializer.is_valid():

    #         serializer.save(user=request.user)
    #         return Response({'status': 'success',
    #                         'message': 'data saved successfully'} , status = 200)

    #     else:
    #         key, value = list(serializer.errors.items())[0]
    #         error_message = key+" , "+value[0]
    #         return Response({'message': error_message,
    #                         'status': 'error'}, status=400)

    def patch(self, request, caseNumber):
        try:
            instance = CaseSheet.objects.get(caseNumber__caseNumber=caseNumber)
        except:
            return Response({'status': 'error',
                            'message': 'caseNumber ID is not found'}, status=400)

        caseNumber = caseNumber.replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['admitCaseClinicalSheet', 'icp', 'medicationChart', 'initialSheet', 'tprChart',
                      'vitalChart', 'losStayChart', 'caseSheetOtherDocuments']:
            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 2000 * 1024:
                        return Response({'status': 'error',
                                        'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
                                        status=400)
                    pdf = PDFGenerator(images, caseNumber, field)
                    try:
                        # check if a zip file already exists for this field
                        if getattr(instance, field):
                            zip_file_instance = getattr(instance, field)
                            if zip_file_instance:
                                with zipfile.ZipFile(zip_file_instance.path, 'a') as zip_file:
                                    pdf_names = zip_file.namelist()
                                    existing_pdf_count = sum(
                                        [1 for pdf_name in pdf_names if pdf_name.startswith(f"{caseNumber}_")])
                                    # set the new PDF name
                                    new_pdf_name = f"{caseNumber}_{date}_{field}_{existing_pdf_count + 1}.pdf"
                                    zip_file.writestr(new_pdf_name, pdf)
                            else:
                                new_file = File(
                                    io.BytesIO(), name=f"{caseNumber}_{date}_{field}.zip")
                                setattr(instance, field, new_file)
                                with zipfile.ZipFile(new_file, 'w') as zip_file:
                                    zip_file.writestr(
                                        f"{caseNumber}_{date}_{field}.pdf", pdf)
                        else:
                            new_file = File(
                                io.BytesIO(), name=f"{caseNumber}_{date}_{field}.zip")
                            setattr(instance, field, new_file)
                            with zipfile.ZipFile(new_file, 'w') as zip_file:
                                zip_file.writestr(
                                    f"{caseNumber}_{date}_{field}.pdf", pdf)
                    except Exception as e:
                        return Response({'status': 'error',
                                        'message': 'Invalid zip file'}, status=400)

                files[field] = getattr(instance, field)
            elif getattr(instance, field):
                files[field] = getattr(instance, field)

        serializer = self.get_serializer(
            instance, data={**files}, partial=True)

        if serializer.is_valid():
            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data saved successfully'}, status=200)
        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key + " , " + value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)

    def delete(self, request, caseNumber):
        # try:
            instance = CaseSheet.objects.get(caseNumber__caseNumber=caseNumber)
            if instance.admitCaseClinicalSheet:
               instance.admitCaseClinicalSheet.delete()
            instance.delete()
            return Response({'status': 'success',
                             'message': 'Deleted successfully'}, status=200)
        # except:
            
            # return Response({'status': 'error',
            #                  'message': 'caseNumber ID is not found'}, status=400)


class LabtestPostForm(generics.GenericAPIView):
    serializer_class = LabTestSerializer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request):
    
    
        caseNumber = request.data['caseNumber'].replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['microbiology', 'biochemistry', 'pathology', 'serologyInvestigation', 'labTestOtherDocuments']:

            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 2000 * 1024:
                        return Response({'status': 'error',
                                         'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
                                        status=400)
                pdf = PDFGenerator(images, caseNumber, field)
                files[field] = SimpleUploadedFile(
                    f"{caseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

        serializer = LabTestSerializer(data={
            'caseNumber': request.data['caseNumber'],
            **files})

        if serializer.is_valid():
            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data saved successfully'})

        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)


class LabtestupdateForm(generics.GenericAPIView):
    serializer_class = LabTestUpdateSerializer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    # def put(self, request , caseNumber):
    #     try:
    #         instance = LabTest.objects.get(caseNumber__caseNumber = caseNumber)
    #     except:
    #         return Response({'status': 'error',
    #                          'message' : 'caseNumber ID is not found'}, status=400)

    #     caseNumber =caseNumber.replace('/', '')
    #     date = datetime.datetime.today().date()
    #     files = {}
    #     for field in ['microbiology', 'biochemistry', 'pathology', 'serologyInvestigation', 'labTestOtherDocuments']:

    #         images = request.FILES.getlist(field)
    #         if images:
    #             for image in images:
    #                     if image.size > 2000 * 1024:
    #                         return Response({'status': 'error',
    #                                          'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
    #                                          status=400 )
    #             pdf = PDFGenerator(images, caseNumber, field)
    #             files[field] = SimpleUploadedFile(
    #                 f"{caseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

    #     serializer = self.get_serializer(instance , data={
    #         'microbiology': files.get('microbiology', None),
    #         'biochemistry': files.get('biochemistry', None),
    #         'pathology': files.get('pathology', None),
    #         'serologyInvestigation': files.get('serologyInvestigation', None),
    #         'labTestOtherDocuments': files.get('labTestOtherDocuments', None)
    #         } , partial = False)

    #     if serializer.is_valid():

    #         serializer.save(user=request.user)
    #         return Response({'status': 'success',
    #                         'message': 'data Updated successfully'})

    #     else:
    #         key, value = list(serializer.errors.items())[0]
    #         error_message = key+" , "+value[0]
    #         return Response({'message': error_message,
    #                         'status': 'error'}, status=400)

    def patch(self, request, caseNumber):
        try:
            instance = LabTest.objects.get(caseNumber__caseNumber=caseNumber)
        except:
            return Response({'status': 'error',
                             'message': 'caseNumber ID is not found'}, status=400)

        caseNumber = caseNumber.replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['microbiology', 'biochemistry', 'pathology', 'serologyInvestigation', 'labTestOtherDocuments']:

            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 2000 * 1024:
                        return Response({'status': 'error',
                                        'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
                                        status=400)
                    pdf = PDFGenerator(images, caseNumber, field)
                    try:
                        # check if a zip file already exists for this field
                        if getattr(instance, field):
                            zip_file_instance = getattr(instance, field)
                            if zip_file_instance:
                                with zipfile.ZipFile(zip_file_instance.path, 'a') as zip_file:
                                    pdf_names = zip_file.namelist()
                                    existing_pdf_count = sum(
                                        [1 for pdf_name in pdf_names if pdf_name.startswith(f"{caseNumber}_")])
                                    # set the new PDF name
                                    new_pdf_name = f"{caseNumber}_{date}_{field}_{existing_pdf_count + 1}.pdf"
                                    zip_file.writestr(new_pdf_name, pdf)
                            else:
                                new_file = File(
                                    io.BytesIO(), name=f"{caseNumber}_{date}_{field}.zip")
                                setattr(instance, field, new_file)
                                with zipfile.ZipFile(new_file, 'w') as zip_file:
                                    zip_file.writestr(
                                        f"{caseNumber}_{date}_{field}.pdf", pdf)
                        else:
                            new_file = File(
                                io.BytesIO(), name=f"{caseNumber}_{date}_{field}.zip")
                            setattr(instance, field, new_file)
                            with zipfile.ZipFile(new_file, 'w') as zip_file:
                                zip_file.writestr(
                                    f"{caseNumber}_{date}_{field}.pdf", pdf)
                    except Exception as e:
                        return Response({'status': 'error',
                                        'message': 'Invalid zip file'}, status=400)

                files[field] = getattr(instance, field)
            elif getattr(instance, field):
                files[field] = getattr(instance, field)

        serializer = self.get_serializer(instance, data={
            **files}, partial=True)

        if serializer.is_valid():

            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data Updated successfully'})

        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)

    def delete(self, request, caseNumber):
        try:
            instance = LabTest.objects.get(caseNumber__caseNumber=caseNumber)
            instance.delete()
            return Response({'status': 'success',
                             'message': 'Deleted successfully'}, status=200)
        except:
            return Response({'status': 'error',
                             'message': 'caseNumber ID is not found'}, status=400)


class ReportsPostForm(generics.GenericAPIView):

    serializer_class = ReportsSerializer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request):

        caseNumber = request.data['caseNumber'].replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['otprocedureSheets', 'anaesthesiaNotes', 'radiology' ,'reportsOtherDocuments']:

            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 2000 * 1024:
                        return Response({'status': 'error',
                                         'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
                                        status=400)
                pdf = PDFGenerator(images, caseNumber, field)
                files[field] = SimpleUploadedFile(
                    f"{caseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

        # radiology_files = request.FILES.getlist('radiology')
        # for i in radiology_files and radiology_files:
        #     if i.name.endswith('.wrf'):
                
        #         zip_buffer = io.BytesIO()
        #         with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        #             for file in radiology_files:
        #                 # Save the original file to disk temporarily
        #                 file_path = os.path.join(
        #                     os.path.dirname(file.name), file.name)
        #                 with open(file_path, 'wb+') as destination:
        #                     for chunk in file.chunks():
        #                         destination.write(chunk)
        #                 # Add the file to the ZIP archive
        #                 zip_file.write(file_path, os.path.basename(file_path))
        #                 # Remove the temporary file from disk
        #                 os.remove(file_path)
        #         # Create a SimpleUploadedFile object from the ZIP buffer
        #         zip_buffer.seek(0)
        #         files['radiology'] = SimpleUploadedFile(f"{caseNumber}_{datetime.datetime.today().date()}_radiology.zip",
        #                                                 zip_buffer.read(),
        #                                                 content_type='application/zip')
        #     else:
        #         return Response({'status': 'error',
        #                         'message': 'Only .wrf files are allowed to be uploaded in radiology Reports'}, status=400)

        serializer = self.get_serializer(data={
            'caseNumber': request.data['caseNumber'],
            'radiology': files.get('radiology', None),
            'otprocedureSheets': files.get('otprocedureSheets', None),
            'anaesthesiaNotes': files.get('anaesthesiaNotes', None),
            'reportsOtherDocuments': files.get('reportsOtherDocuments', None),
        })
        if serializer.is_valid():
            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data saved successfully'})
        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)


class ReportsUpdateForm(generics.GenericAPIView):
    serializer_class = ReportsUpdateSerializer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    # def put(self , request , caseNumber):
    #     try:
    #         instance = Reports.objects.get(caseNumber__caseNumber = caseNumber)
    #     except:
    #         return Response({'status': 'error',
    #                          'message' : 'caseNumber ID is not found'}, status=400)

    #     caseNumber =caseNumber.replace('/', '')
    #     date = datetime.datetime.today().date()
    #     files = {}
    #     for field in ['otprocedureSheets', 'anaesthesiaNotes', 'reportsOtherDocuments']:

    #         images = request.FILES.getlist(field)
    #         if images:
    #             for image in images:
    #                     if image.size > 2000 * 1024:
    #                         return Response({'status': 'error',
    #                                          'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
    #                                          status=400 )
    #             pdf = PDFGenerator(images, caseNumber, field)
    #             files[field] = SimpleUploadedFile(
    #                 f"{caseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

    #     radiology_files = request.FILES.getlist('radiology')
    #     if radiology_files:
    #         for i in radiology_files:
    #             if i.name.endswith('.wrf'):
    #                 zip_buffer = io.BytesIO()
    #                 with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
    #                     for file in radiology_files:
    #                         # Save the original file to disk temporarily
    #                         file_path = os.path.join(
    #                             os.path.dirname(file.name), file.name)
    #                         with open(file_path, 'wb+') as destination:
    #                             for chunk in file.chunks():
    #                                 destination.write(chunk)
    #                         # Add the file to the ZIP archive
    #                         zip_file.write(file_path, os.path.basename(file_path))
    #                         # Remove the temporary file from disk
    #                         os.remove(file_path)
    #                 # Create a SimpleUploadedFile object from the ZIP buffer
    #                 zip_buffer.seek(0)
    #                 files['radiology'] = SimpleUploadedFile(f"{caseNumber}_{date}_radiology.zip",
    #                                                         zip_buffer.read(), content_type='application/zip')
    #             else:
    #                 return Response({"status" : "error", "message": 'Only .wrf files are allowed to be uploaded in radiology Reports'}, status= 400)
    #     else:
    #         files['radiology'] = None

    #     serializer = self.get_serializer(instance,data={
    #         'radiology': files.get('radiology', None),
    #         'otprocedureSheets': files.get('otprocedureSheets', None),
    #         'anaesthesiaNotes': files.get('anaesthesiaNotes', None),
    #         'reportsOtherDocuments': files.get('reportsOtherDocuments', None),} , partial = False)

    #     if serializer.is_valid():
    #         serializer.save(user=request.user)
    #         return Response({'status': 'success',
    #                         'message': 'data saved successfully'})
    #     else:
    #         key, value = list(serializer.errors.items())[0]
    #         error_message = key+" , "+value[0]
    #         return Response({'message': error_message,
    #                         'status': 'error'}, status=400)

    def patch(self, request, caseNumber):
        try:
            instance = Reports.objects.get(caseNumber__caseNumber=caseNumber)
        except:
            return Response({'status': 'error',
                             'message': 'caseNumber ID is not found'}, status=400)

        caseNumber = caseNumber.replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['otprocedureSheets', 'anaesthesiaNotes', 'radiology' , 'reportsOtherDocuments']:

            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 2000 * 1024:
                        return Response({'status': 'error',
                                        'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
                                        status=400)
                    pdf = PDFGenerator(images, caseNumber, field)
                    try:
                        # check if a zip file already exists for this field
                        if getattr(instance, field):
                            zip_file_instance = getattr(instance, field)
                            if zip_file_instance:
                                with zipfile.ZipFile(zip_file_instance.path, 'a') as zip_file:
                                    pdf_names = zip_file.namelist()
                                    existing_pdf_count = sum(
                                        [1 for pdf_name in pdf_names if pdf_name.startswith(f"{caseNumber}_")])
                                    # set the new PDF name
                                    new_pdf_name = f"{caseNumber}_{date}_{field}_{existing_pdf_count + 1}.pdf"
                                    zip_file.writestr(new_pdf_name, pdf)
                            else:
                                new_file = File(
                                    io.BytesIO(), name=f"{caseNumber}_{date}_{field}.zip")
                                setattr(instance, field, new_file)
                                with zipfile.ZipFile(new_file, 'w') as zip_file:
                                    zip_file.writestr(
                                        f"{caseNumber}_{date}_{field}.pdf", pdf)
                        else:
                            new_file = File(
                                io.BytesIO(), name=f"{caseNumber}_{date}_{field}.zip")
                            setattr(instance, field, new_file)
                            with zipfile.ZipFile(new_file, 'w') as zip_file:
                                zip_file.writestr(
                                    f"{caseNumber}_{date}_{field}.pdf", pdf)
                    except Exception as e:
                        return Response({'status': 'error',
                                        'message': 'Invalid zip file'}, status=400)

                files[field] = getattr(instance, field)
            elif getattr(instance, field):
                files[field] = getattr(instance, field)

        # radiology_files = request.FILES.getlist('radiology')
        # if radiology_files:
        #     for i in radiology_files:
        #         if i.name.endswith('.wrf'):
        #             zip_buffer = io.BytesIO()
        #             with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        #                 for file in radiology_files:
        #                     # Save the original file to disk temporarily
        #                     file_path = os.path.join(
        #                         os.path.dirname(file.name), file.name)
        #                     with open(file_path, 'wb+') as destination:
        #                         for chunk in file.chunks():
        #                             destination.write(chunk)
        #                     # Add the file to the ZIP archive
        #                     zip_file.write(
        #                         file_path, os.path.basename(file_path))
        #                     # Remove the temporary file from disk
        #                     os.remove(file_path)
        #             # Create a SimpleUploadedFile object from the ZIP buffer
        #             zip_buffer.seek(0)
        #             files['radiology'] = SimpleUploadedFile(f"{caseNumber}_{date}_radiology.zip",
        #                                                     zip_buffer.read(), content_type='application/zip')
        #         else:
        #             return Response({"status": "error", "message": 'Only .wrf files are allowed to be uploaded in radiology Reports'}, status=400)
        # else:
        #     pass

        serializer = self.get_serializer(
            instance, data={**files}, partial=False)

        if serializer.is_valid():

            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data updated successfully'})
        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)

    def delete(self, request, caseNumber):
        try:
            instance = Reports.objects.get(caseNumber__caseNumber=caseNumber)
            instance.delete()
            return Response({'status': 'success',
                             'message': 'Deleted successfully'}, status=200)
        except:
            return Response({'status': 'error',
                             'message': 'caseNumber ID is not found'}, status=400)


class DischargesummuryPostAPI(generics.GenericAPIView):

    serializer_class = DischargeSummarySerialzer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):

        caseNumber = request.data['caseNumber'].replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['dischargeSummaryDocument', 'dischargeSummaryOtherDocuments']:

            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 2000 * 1024:
                        return Response({'status': 'error',
                                         'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
                                        status=400)

                pdf = PDFGenerator(images, caseNumber, field)
                files[field] = SimpleUploadedFile(
                    f"{caseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

        serializer = self.get_serializer(data={
            'caseNumber': request.data['caseNumber'],
            'dischargeDate': request.data['dischargeDate'],
            'dischargeType': request.data['dischargeType'],
            'dischargeSummaryDocument': files.get('dischargeSummaryDocument', None),
            'OtherDocuments': files.get('OtherDocuments', None)
        })

        if serializer.is_valid():
            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data saved successfully'})

        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)


class DischargesummuryUpdateAPI(generics.GenericAPIView):
    serializer_class = DischargeSummaryUpdateSerialzer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    # def put(self, request, caseNumber):
    #     try:
    #         instance = DischargeSummary.objects.get(caseNumber__caseNumber = caseNumber)
    #     except:
    #         return Response({'status': 'error',
    #                          'message' : 'caseNumber ID is not found'}, status=400)

    #     caseNumber =caseNumber.replace('/', '')
    #     date = datetime.datetime.today().date()
    #     files = {}
    #     for field in ['dischargeSummaryDocument', 'dischargeSummaryOtherDocuments']:

    #         images = request.FILES.getlist(field)
    #         if images:
    #             for image in images:
    #                     if image.size > 2000 * 1024:
    #                         return Response({'status': 'error',
    #                                          'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
    #                                          status=400 )
    #             pdf = PDFGenerator(images, caseNumber, field)
    #             files[field] = SimpleUploadedFile(
    #                 f"{caseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

    #     serializer = self.get_serializer(instance , data={
    #         'dischargeDate': request.data['dischargeDate'],
    #         'dischargeType': request.data['dischargeType'],
    #         'dischargeSummaryDocument': files.get('dischargeSummaryDocument', None),
    #         'dischargeSummaryOtherDocuments': files.get('dischargeSummaryOtherDocuments', None)
    #     })

    #     if serializer.is_valid():
    #         serializer.save(user=request.user)
    #         return Response({'status': 'success',
    #                         'message': 'data saved successfully'})

    #     else:
    #         key, value = list(serializer.errors.items())[0]
    #         error_message = key+" , "+value[0]
    #         return Response({'message': error_message,
    #                         'status': 'error'}, status=400)

    def patch(self, request, caseNumber):
        try:
            instance = DischargeSummary.objects.get(
                caseNumber__caseNumber=caseNumber)
        except:
            return Response({'status': 'error',
                             'message': 'caseNumber ID is not found'}, status=400)

        caseNumber = caseNumber.replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['dischargeSummaryDocument', 'dischargeSummaryOtherDocuments']:
            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 2000 * 1024:
                        return Response({'status': 'error',
                                        'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
                                        status=400)
                    pdf = PDFGenerator(images, caseNumber, field)
                    try:
                        # check if a zip file already exists for this field
                        if getattr(instance, field):
                            zip_file_instance = getattr(instance, field)
                            if zip_file_instance:
                                with zipfile.ZipFile(zip_file_instance.path, 'a') as zip_file:
                                    pdf_names = zip_file.namelist()
                                    existing_pdf_count = sum(
                                        [1 for pdf_name in pdf_names if pdf_name.startswith(f"{caseNumber}_")])
                                    # set the new PDF name
                                    new_pdf_name = f"{caseNumber}_{date}_{field}_{existing_pdf_count + 1}.pdf"
                                    zip_file.writestr(new_pdf_name, pdf)
                            else:
                                new_file = File(
                                    io.BytesIO(), name=f"{caseNumber}_{date}_{field}.zip")
                                setattr(instance, field, new_file)
                                with zipfile.ZipFile(new_file, 'w') as zip_file:
                                    zip_file.writestr(
                                        f"{caseNumber}_{date}_{field}.pdf", pdf)
                        else:
                            new_file = File(
                                io.BytesIO(), name=f"{caseNumber}_{date}_{field}.zip")
                            setattr(instance, field, new_file)
                            with zipfile.ZipFile(new_file, 'w') as zip_file:
                                zip_file.writestr(
                                    f"{caseNumber}_{date}_{field}.pdf", pdf)
                    except Exception as e:
                        return Response({'status': 'error',
                                        'message': 'Invalid zip file'}, status=400)

                files[field] = getattr(instance, field)
            elif getattr(instance, field):
                files[field] = getattr(instance, field)

        new_data = {**files}
        if 'dischargeDate' in request.data:
            new_data['dischargeDate'] = request.data['dischargeDate']

        if 'dischargeType' in request.data:
            new_data['dischargeType'] = request.data['dischargeType']

        serializer = self.get_serializer(instance, data=new_data, partial=True)
        if serializer.is_valid():

            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data saved successfully'})

        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)

    def delete(self, request, caseNumber):
        try:
            instance = DischargeSummary.objects.get(
                caseNumber__caseNumber=caseNumber)
            instance.delete()
            return Response({'status': 'success',
                             'message': 'Deleted successfully'}, status=200)
        except:
            return Response({'status': 'error',
                             'message': 'caseNumber ID is not found'}, status=400)


class DeathSummuryPostAPI(generics.GenericAPIView):
    serializer_class = DeathSummarySerialzer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        caseNumber = request.data['caseNumber'].replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['mortalityAudit', 'deathCertificate', 'deathSummaryOtherDocuments']:

            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 2000 * 1024:
                        return Response({'status': 'error',
                                         'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
                                        status=400)
                pdf = PDFGenerator(images, caseNumber, field)
                files[field] = SimpleUploadedFile(
                    f"{caseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

        serializer = self.get_serializer(data={
            'caseNumber': request.data['caseNumber'],
            'mortalityAudit': files.get('mortalityAudit', None),
            'deathCertificate': files.get('deathCertificate', None),
            'deathSummaryOtherDocuments': files.get('deathSummaryOtherDocuments', None)
        })

        if serializer.is_valid():

            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data saved successfully'})

        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)


class DeathSummuryUpdateAPI(generics.GenericAPIView):
    serializer_class = DeathSummaryUpdateSerialzer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    # def put (self, request , caseNumber):
    #     try:
    #         instance = DeathSummary.objects.get(caseNumber__caseNumber = caseNumber)
    #     except:
    #         return Response({'status': 'error',
    #                          'message' : 'caseNumber ID is not found'}, status=400)

    #     caseNumber =caseNumber.replace('/', '')
    #     date = datetime.datetime.today().date()
    #     files = {}
    #     for field in ['mortalityAudit', 'deathCertificate', 'deathSummaryOtherDocuments']:

    #         images = request.FILES.getlist(field)
    #         if images:
    #             for image in images:
    #                     if image.size > 2000 * 1024:
    #                         return Response({'status': 'error',
    #                                          'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
    #                                          status=400 )
    #             pdf = PDFGenerator(images, caseNumber, field)
    #             files[field] = SimpleUploadedFile(
    #                 f"{caseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

    #     serializer = self.get_serializer(instance ,  data={
    #         'mortalityAudit': files.get('mortalityAudit', None),
    #         'deathCertificate': files.get('deathCertificate', None),
    #         'deathSummaryOtherDocuments': files.get('deathSummaryOtherDocuments', None)} , partial = False)

    #     if serializer.is_valid():
    #         serializer.save(user=request.user)
    #         return Response({'status': 'success',
    #                         'message': 'data updated successfully'})

    #     else:
    #         key, value = list(serializer.errors.items())[0]
    #         error_message = key+" , "+value[0]
    #         return Response({'message': error_message,
    #                         'status': 'error'}, status=400)

    def patch(self, request, caseNumber):
        try:
            instance = DeathSummary.objects.get(
                caseNumber__caseNumber=caseNumber)
        except:
            return Response({'status': 'error',
                             'message': 'caseNumber ID is not found'}, status=400)

        caseNumber = caseNumber.replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['mortalityAudit', 'deathCertificate', 'deathSummaryOtherDocuments']:

            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 2000 * 1024:
                        return Response({'status': 'error',
                                        'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
                                        status=400)
                    pdf = PDFGenerator(images, caseNumber, field)
                    try:
                        # check if a zip file already exists for this field
                        if getattr(instance, field):
                            zip_file_instance = getattr(instance, field)
                            if zip_file_instance:
                                with zipfile.ZipFile(zip_file_instance.path, 'a') as zip_file:
                                    pdf_names = zip_file.namelist()
                                    existing_pdf_count = sum(
                                        [1 for pdf_name in pdf_names if pdf_name.startswith(f"{caseNumber}_")])
                                    # set the new PDF name
                                    new_pdf_name = f"{caseNumber}_{date}_{field}_{existing_pdf_count + 1}.pdf"
                                    zip_file.writestr(new_pdf_name, pdf)
                            else:
                                new_file = File(
                                    io.BytesIO(), name=f"{caseNumber}_{date}_{field}.zip")
                                setattr(instance, field, new_file)
                                with zipfile.ZipFile(new_file, 'w') as zip_file:
                                    zip_file.writestr(
                                        f"{caseNumber}_{date}_{field}.pdf", pdf)
                        else:
                            new_file = File(
                                io.BytesIO(), name=f"{caseNumber}_{date}_{field}.zip")
                            setattr(instance, field, new_file)
                            with zipfile.ZipFile(new_file, 'w') as zip_file:
                                zip_file.writestr(
                                    f"{caseNumber}_{date}_{field}.pdf", pdf)
                    except Exception as e:
                        return Response({'status': 'error',
                                        'message': 'Invalid zip file'}, status=400)

                files[field] = getattr(instance, field)
            elif getattr(instance, field):
                files[field] = getattr(instance, field)

        serializer = self.get_serializer(
            instance,  data={**files}, partial=False)

        if serializer.is_valid():
            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data updated successfully'})

        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)

    def delete(self, request, caseNumber):
        try:
            instance = DeathSummary.objects.get(
                caseNumber__caseNumber=caseNumber)
            instance.delete()
            return Response({'status': 'success',
                             'message': 'Deleted successfully'}, status=200)
        except:
            return Response({'status': 'error',
                             'message': 'caseNumber ID is not found'}, status=400)


class BloodDocumentsPostAPi(generics.GenericAPIView):

    serializer_class = BloodDocumentsSerializer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):

        caseNumber = request.data['caseNumber'].replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['bloodTransfusion', 'btSticker', 'crossMatchReport', 'bloodOtherDocuments']:

            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 2000 * 1024:
                        return Response({'status': 'error',
                                         'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
                                        status=400)

                pdf = PDFGenerator(images, caseNumber, field)
                files[field] = SimpleUploadedFile(
                    f"{caseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

        serializer = self.get_serializer(data={
            'caseNumber': request.data['caseNumber'],
            'bloodTransfusion': files.get('bloodTransfusion', None),
            'btSticker': files.get('btSticker', None),
            'crossMatchReport': files.get('crossMatchReport', None),
            'bloodOtherDocuments': files.get('bloodOtherDocuments', None),
        })

        if serializer.is_valid():
            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data saved successfully'})

        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)


class BloodDocumentsUpdateApi(generics.GenericAPIView):
    serializer_class = BloodDocumentsUpdateSerializer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    # def put (self, request , caseNumber):
    #     try:
    #         instance = BloodDocuments.objects.get(caseNumber__caseNumber = caseNumber)
    #     except:
    #         return Response({'status': 'error',
    #                          'message' : 'caseNumber ID is not found'}, status=400)

    #     caseNumber =caseNumber.replace('/', '')
    #     date = datetime.datetime.today().date()
    #     files = {}
    #     for field in ['bloodTransfusion', 'btSticker', 'crossMatchReport', 'bloodOtherDocuments']:

    #         images = request.FILES.getlist(field)
    #         if images:
    #             for image in images:
    #                     if image.size > 2000 * 1024:
    #                         return Response({'status': 'error',
    #                                          'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
    #                                          status=400 )
    #             pdf = PDFGenerator(images, caseNumber, field)
    #             files[field] = SimpleUploadedFile(
    #                 f"{caseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

    #     serializer = self.get_serializer(instance , data={
    #         'bloodTransfusion': files.get('bloodTransfusion', None),
    #         'btSticker': files.get('btSticker', None),
    #         'crossMatchReport': files.get('crossMatchReport', None),
    #         'bloodOtherDocuments': files.get('bloodOtherDocuments', None),
    #     } , partial = False)

    #     if serializer.is_valid():
    #         serializer.save(user=request.user)
    #         return Response({'status': 'success',
    #                         'message': 'data saved successfully'})

    #     else:
    #         key, value = list(serializer.errors.items())[0]
    #         error_message = key+" , "+value[0]
    #         return Response({'message': error_message,
    #                         'status': 'error'}, status=400)

    def patch(self, request, caseNumber):
        try:
            instance = BloodDocuments.objects.get(
                caseNumber__caseNumber=caseNumber)
        except:
            return Response({'status': 'error',
                             'message': 'caseNumber ID is not found'}, status=400)

        caseNumber = caseNumber.replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['bloodTransfusion', 'btSticker', 'crossMatchReport', 'bloodOtherDocuments']:
            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 2000 * 1024:
                        return Response({'status': 'error',
                                        'message': 'Each Image size should be less than 1.99 MB - {name}'.format(name=image.name)},
                                        status=400)
                    pdf = PDFGenerator(images, caseNumber, field)
                    try:
                        # check if a zip file already exists for this field
                        if getattr(instance, field):
                            zip_file_instance = getattr(instance, field)
                            if zip_file_instance:
                                with zipfile.ZipFile(zip_file_instance.path, 'a') as zip_file:
                                    pdf_names = zip_file.namelist()
                                    existing_pdf_count = sum(
                                        [1 for pdf_name in pdf_names if pdf_name.startswith(f"{caseNumber}_")])
                                    # set the new PDF name
                                    new_pdf_name = f"{caseNumber}_{date}_{field}_{existing_pdf_count + 1}.pdf"
                                    zip_file.writestr(new_pdf_name, pdf)
                            else:
                                new_file = File(
                                    io.BytesIO(), name=f"{caseNumber}_{date}_{field}.zip")
                                setattr(instance, field, new_file)
                                with zipfile.ZipFile(new_file, 'w') as zip_file:
                                    zip_file.writestr(
                                        f"{caseNumber}_{date}_{field}.pdf", pdf)
                        else:
                            new_file = File(
                                io.BytesIO(), name=f"{caseNumber}_{date}_{field}.zip")
                            setattr(instance, field, new_file)
                            with zipfile.ZipFile(new_file, 'w') as zip_file:
                                zip_file.writestr(
                                    f"{caseNumber}_{date}_{field}.pdf", pdf)
                    except Exception as e:
                        return Response({'status': 'error',
                                        'message': 'Invalid zip file'}, status=400)

                files[field] = getattr(instance, field)
            elif getattr(instance, field):
                files[field] = getattr(instance, field)

        serializer = self.get_serializer(
            instance, data={**files}, partial=True)

        if serializer.is_valid():
            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data saved successfully'})

        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)

    def delete(self, request, caseNumber):
        try:
            instance = BloodDocuments.objects.get(
                caseNumber__caseNumber=caseNumber)
            instance.delete()
            return Response({'status': 'success',
                             'message': 'Deleted successfully'}, status=200)
        except:
            return Response({'status': 'error',
                             'message': 'caseNumber ID is not found'}, status=400)


class ClaimManagmentStatusPostAPI(generics.GenericAPIView):
    serializer_class = ClaimManagemntStatusSerialzer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        data = request.data
        serializer = self.get_serializer(data=data)
        if serializer.is_valid():
            serializer.save(user=request.user)
            return Response({'status': 'success',
                             'message': 'data saved successfully'}, status=201)
        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)

class GetClaimCountList(generics.ListAPIView):
    pagination_class = LimitOffsetPagination
    serializer_class = GetLinkedCaseNumberListserialzier
    filter_backends = [SearchFilter]
    search_fileds = ['caseNumber']
    queryset = PreAuthLinkcaseNumber.objects.all().order_by('-date_modified')


    def get(self , request):
        page = self.paginate_queryset(self.get_queryset())
        serializer = self.get_serializer(self.get_queryset() , many = True).data
        if page is not None:
            serializer = self.get_serializer(page, many=True).data
            return self.get_paginated_response({'status': 'success',
                                                'message': 'Data fetched successfully',
                                                'data': serializer})
        
        serializer = self.get_serializer(self.get_queryset(), many=True).data
        return Response({'status': 'success',
                        'message': 'Data fetched successfully',
                         'data': serializer})

class DownloadClaimDocuments(APIView):
    def get(self, request, caseNumber, choice=None):
        Attribute = []
        model_fields = OrderedDict([
            (CaseSheet, 'CaseSheet_query'),
            (LabTest, 'LabTest_query'),
            (Reports, 'Reports_query'),
            (DischargeSummary, 'DischargeSummary_query'),
            (DeathSummary, 'DeathSummary_query'),
            (BloodDocuments, 'BloodDocuments_query'),
        ])
        for model, field_name in model_fields.items():
            try:
                query = model.objects.filter(caseNumber__caseNumber=caseNumber).latest('date_modified')
                Attribute.append(query)
            except model.DoesNotExist:
                pass

        if not Attribute:
            return Response({'status': 'error',
                            'message': f'No Documents Found for This {caseNumber}'},
                            status=400)
        caseNumber = caseNumber.replace('/', '')
        date = datetime.datetime.today().date()
        

        response = HttpResponse(content_type='application/zip'  )
        response['Content-Disposition'] = 'attachment; filename="{name}_{date}_ClaimManagment_Documents.zip"'.format(name=caseNumber, date=date)

        zip_file = zipfile.ZipFile(response, 'w', zipfile.ZIP_DEFLATED)

        for i in Attribute:
            for field_name in ['admitCaseClinicalSheet', 'icp', 'medicationChart', 'initialSheet', 'tprChart', 'vitalChart', 'losStayChart', 'caseSheetOtherDocuments',
                               'microbiology', 'biochemistry', 'pathology', 'serologyInvestigation', 'labTestOtherDocuments',
                               'radiology', 'otprocedureSheets', 'anaesthesiaNotes', 'reportsOtherDocuments',
                               'dischargeSummaryDocument', 'dischargeSummaryOtherDocuments',
                               'mortalityAudit', 'deathCertificate', 'deathSummaryOtherDocuments',
                               'bloodTransfusion', 'btSticker', 'crossMatchReport', 'bloodOtherDocuments']:
                if hasattr(i, field_name):
                    field = getattr(i, field_name)
                    try:
                        if field and field.file:
                            with zipfile.ZipFile(field.file, 'r') as zip_ref:
                                for zip_info in zip_ref.infolist():
                                    zip_info.filename = os.path.basename(
                                        zip_info.filename)
                                    zip_file.writestr(
                                        zip_info, zip_ref.read(zip_info.filename))
                    except:
                        continue

        zip_file.close()
        if choice == 'SP':
            share_path = '//192.168.1.7/SharePath/Download'
            filename = response.get('Content-Disposition').split('filename=')[1].replace('"', '')

            zip_file_path = os.path.join(share_path, filename)

            with open(zip_file_path, 'wb') as f:
                f.write(response.content)

            # Return a success response
            logger.info(f'File downloaded successfully - {zip_file_path}')
            return Response({'status': 'success',
                            'message': 'File downloaded and saved to share path successfully'}, status=200)
        elif choice == 'PC':
            return response 
        else:
            return Response({'status': 'error', 'message': 'Invalid choice for download'}, status=400)


# inserting the data in to database for .xls file
class DumpExcelInsertxls(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = DumpExcelSerializer
    permission_classes = [IsAuthenticated]

    def post(self, request, format=None):
        try:
            if 'excel_file' not in request.FILES:
                return Response({"status": "error", "message": "No file uploaded."}, status=400)

            excel_file = request.FILES["excel_file"]
            if excel_file.size == 0 or not excel_file.name.endswith(".xls"):
                return Response({"status": "error", "message": "Only .xls file is supported."}, status=400)

            try:
                wb = xlrd.open_workbook(file_contents=excel_file.read())
                sheet_name = wb.sheet_names()[0]
                worksheet = wb.sheet_by_name(sheet_name)

                data_list = []
                for row_idx in range(3, worksheet.nrows):
                    row = worksheet.row_values(row_idx)
                    # Replace "NA", "-NA-", and "na" with None
                    row = [None if x in ["NA", "-NA-", "na"] else x for x in row]
                    datetime_fields = [16, 17, 19, 21, 22, 23, 24, 25, 26, 34]
                    for field_idx in datetime_fields:
                        if row[field_idx]:
                            try:
                                row[field_idx] = datetime.datetime.strptime(
                                    row[field_idx], "%d/%m/%Y %I:%M:%S %p").strftime('%Y-%m-%d %H:%M:%S')
                            except ValueError:
                                row[field_idx] = None
                        else:
                            row[field_idx] = None
                    date_fields = [15]
                    for field_idx in date_fields:
                        if row[field_idx]:
                            try:
                                row[field_idx] = datetime.datetime.strptime(
                                    row[field_idx], '%d/%m/%Y').strftime('%Y-%m-%d')
                            except ValueError:
                                row[field_idx] = ""
                        else:
                            row[field_idx] = ""
                    try:
                        NHPM_creation = PersonalInfo.objects.get(nhmpID = row[3])
                    except:
                        NHPM_creation = PersonalInfo.objects.create(user = request.user, nhmpID = row[3] , gender = row[6] , nameOfPatient = '' )

                        
                    try:
                        existing_data = DumpExcel.objects.get(caseNo_id=row[0])
                    except DumpExcel.DoesNotExist:
                        existing_data = None

                    if existing_data:
                        existing_data.user = request.user
                        existing_data.claimNo = row[1]
                        existing_data.registrationNumber = row[2]
                        existing_data.nhmpID = row[3]
                        existing_data.familyId = row[4]
                        existing_data.patientDistrict = row[5]
                        existing_data.gender = row[6]
                        existing_data.age = row[7]
                        existing_data.categoryDetails = row[8]
                        existing_data.procedureDetails = row[9]
                        existing_data.caseType = row[10]
                        existing_data.caseStatus = row[11]
                        existing_data.hospitalName = row[12]
                        existing_data.hospitalCode = row[13]
                        existing_data.hospitalDistrict = row[14]
                        existing_data.ipRegistrationDate = row[15]
                        existing_data.admissionDate = row[16]
                        existing_data.preauthDate = row[17]
                        existing_data.preauthAmount = row[18]
                        existing_data.preauthApproveDate = row[19]
                        existing_data.preauthApprovedAmount = row[20]
                        existing_data.preauthRejectedDate = row[21]
                        existing_data.surgeryDate = row[22]
                        existing_data.deathDate = row[23]
                        existing_data.dischargeDate = row[24]
                        existing_data.claimSubmittedDate = row[25]
                        existing_data.actualclaimSubmittedDate = row[26]
                        existing_data.claimInitaiatedAmount = row[27]
                        existing_data.cpdApprovedAmount = row[28]
                        existing_data.claimApprovedAmount = row[29]
                        existing_data.assignedFlag = row[30]
                        existing_data.assignedUser = row[31]
                        existing_data.assignedGroup = row[32]
                        existing_data.ipNumber = row [33]
                        existing_data.actualRegistrationDate = row [34]
                    
                        existing_data.save()
                    else:
                        try:
                            get_linked_case_number = PreAuthLinkcaseNumber.objects.get(caseNumber = row[0])
                        except:
                            Linked_caseNumber_creation = PreAuthLinkcaseNumber.objects.create(caseNumber = row[0] , user = request.user)
                        data = DumpExcel(
                        user=request.user,
                        caseNo_id=row[0], claimNo=row[1], registrationNumber=row[2], nhmpID=row[3],
                        familyId=row[4], patientDistrict=row[5], gender=row[6], age=row[7],
                        categoryDetails=row[8], procedureDetails=row[9], caseType=row[10], caseStatus=row[11], hospitalName=row[12],
                        hospitalCode=row[13], hospitalDistrict=row[14], ipRegistrationDate=row[
                            15], admissionDate=row[16], preauthDate=row[17],
                        preauthAmount=row[18], preauthApproveDate=row[19], preauthApprovedAmount=row[20], preauthRejectedDate=row[21],
                        surgeryDate=row[22], deathDate=row[23], dischargeDate=row[
                            24], claimSubmittedDate=row[25], actualclaimSubmittedDate=row[26],
                        claimInitaiatedAmount=row[27], cpdApprovedAmount=row[
                            28], claimApprovedAmount=row[29], assignedFlag=row[30], assignedUser=row[31],
                        assignedGroup=row[32], ipNumber=row[33], actualRegistrationDate=row[34]
                        )
                        data_list.append(data)

                DumpExcel.objects.bulk_create(data_list)
                return Response({"status": "Success",
                        "message": "Successfully Uploaded."})


            except Exception as e:
                error_message = str(e)
                detail_message = error_message.split(":")[-1].strip()
                if detail_message:
                    caseNumber = detail_message.split('(')[2].split(')')[0]
                    return Response({"status": "error", "message": f"Error when attempting to store this Case Number : {caseNumber} , It's because this Case Number not linked to any PreAuth."}, status=400)
        except Exception as e:
            print(str(e))
            return Response({"status": "error", "message": "Please Check your excel File Format and upload it again."}, status=400)


# for inserting the data into database .xlsx file format... No use Right now
class DumpExcelInsertxlsx(generics.GenericAPIView):
    
    parser_classes = [MultiPartParser]
    serializer_class = DumpExcelSerializer

    def post(self, request, format=None):
        # try:
        if 'excel_file' not in request.FILES:
            return Response({"status": "error", "message": "No file uploaded."}, status=400)

        excel_file = request.FILES["excel_file"]
       
        if excel_file.size == 0 or excel_file.name.endswith(".xlsx") != True:
            return Response({"status": "error",
                             "message": "only .xlsx file is supported."},
                            status=400)

        workbook = load_workbook(filename=excel_file)
        sheet_name = workbook.sheetnames[0]
        worksheet = workbook[sheet_name]
        i = 0
        for r in range(1, worksheet.max_row+1):
            for c in range(1, worksheet.max_column+1):
                cell_value = str(worksheet.cell(r, c).value)
                if "-NA-" in cell_value:
                    worksheet.cell(r, c).value = cell_value.replace("-NA-", "")
                    i += 1
                elif cell_value.lower() in ["na", "nan"]:
                    worksheet.cell(r, c).value = ""
                    i += 1

        date_fields = [15]
        datetime_fields = [16, 17, 18]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            print(row)
            for field_idx in datetime_fields:
                if row[field_idx]:
                    try:
                        row[field_idx] = datetime.datetime.strptime(
                            row[field_idx], "%d/%m/%Y %I:%M:%S %p").strftime('%Y-%m-%d %H:%M:%S')
                    except ValueError as e:
                        row[field_idx] = ""

            for field_idx in date_fields:
                if row[field_idx]:
                    try:
                        row[field_idx] = datetime.datetime.strptime(
                            row[field_idx], '%d/%m/%Y').strftime('%Y-%m-%d')
                    except ValueError:
                        row[field_idx] = ""
        # print(row)
        data_list = [ DumpExcel(*row) for row in worksheet.iter_rows(min_row=2, values_only=True)]

        DumpExcel.objects.bulk_create(data_list)
        return Response({"status": "Success",
                        "message": "Successfully Uploaded."})
        # except:
        #     return Response({"status": "error",
        #                     "message": "Something Went Wrong Please Check your Excel File and upload Again."},
        #                     status=400)


class MPClaimPiadPostApi(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = DumpExcelSerializer
    permission_classes = [IsAuthenticated]

    def post(self, request, format=None):
        try:
            if 'excel_file' not in request.FILES:
                return Response({"status": "error", "message": "No file uploaded."}, status=400)

            excel_file = request.FILES["excel_file"]
            if excel_file.size == 0 or excel_file.name.endswith(".xlsx") != True:
                return Response({"status": "error", "message": "only .xlsx file is supported."}, status=400)
            try:
                # Load updated Excel file into openpyxl workbook and create database object
                workbook = load_workbook(filename=excel_file)
                sheet_name = workbook.sheetnames[0]
                worksheet = workbook[sheet_name]
                i = 0
                for r in range(1, worksheet.max_row+1):
                    for c in range(1, worksheet.max_column+1):
                        cell_value = str(worksheet.cell(r, c).value)
                        if "-NA-" in cell_value:
                            worksheet.cell(r, c).value = cell_value.replace("-NA-", "")
                            i += 1
                        elif cell_value.lower() in ["na", "nan"]:
                            worksheet.cell(r, c).value = ""
                            i += 1

                datetime_fields = [3, 25, 27, 28, 30, 31, 33, 36, 37, 38, 39, 41, 42, 43,
                                    44, 57, 59, 61, 63, 65, 71, 73, 74, 80, 85]

                for r in range(2, worksheet.max_row+1):
                    for c in datetime_fields:
                        cell_value = worksheet.cell(row=r, column=c).value
                      
                        if isinstance(cell_value, datetime.datetime):
                            worksheet.cell(row=r, column=c).value = cell_value.strftime(
                                '%Y-%m-%d %H:%M:%S')
                        
                        elif cell_value and cell_value.strip():  # check if value is not None and not empty after stripping whitespaces
                            try:
                                cell_value = datetime.datetime.strptime(cell_value, "%d/%m/%Y%I:%M:%S %p")
                                worksheet.cell(row=r, column=c).value = cell_value.strftime('%Y-%m-%d %H:%M:%S')
                            except ValueError as e:
                                worksheet.cell(row=r, column=c).value = None
                        else:
                            worksheet.cell(row=r, column=c).value = None

                
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    for row in worksheet.iter_rows(min_row=2, values_only=True):
                        case_number = row[0]
                        existing_record = MPClaimPaidExcel.objects.filter(caseNumber_id=case_number).first()
                        try:
                            NHPM_card_id = PersonalInfo.objects.get( nhmpID= row[3])
                        except:
                            create_persnolInfo = PersonalInfo.objects.create( user = request.user ,nhmpID = row[3] , nameOfPatient = row[5]  , gender = row[7])
                        if existing_record:
                            existing_record.user = request.user
                            existing_record.workflowStatus = row[1]
                            existing_record.statusUpdateDate = row[2]
                            existing_record.nhmpID = row[3]
                            existing_record.familyId = row[4]
                            existing_record.patientName = row[5]
                            existing_record.age = row[6]
                            existing_record.gender = row[7]
                            existing_record.patientsHomeDistrict = row[8]
                            existing_record.patientsHomeState = row[9]
                            existing_record.patientPhoneNo = row[10]
                            existing_record.address = row[11]
                            existing_record.communicationAddress = row[12]
                            existing_record.communicationContactNo = row[13]
                            existing_record.communicationVillage = row[14]
                            existing_record.communicationMandal = row[15]
                            existing_record.communicationDistrict = row[16]
                            existing_record.communicationState = row[17]
                            existing_record.speciality = row[18]
                            existing_record.patientIPOP = row[19]
                            existing_record.procedureCode = row[20]
                            existing_record.procedure = row[21]
                            existing_record.procedureAutoApprove = row[22]
                            existing_record.medicalOrSurgery = row[23]
                            existing_record.ipRegistrationDate = row[24]
                            existing_record.hospitalName = row[25]
                            existing_record.admissionDate = row[26]
                            existing_record.preAuthInitiationDate = row[27]
                            existing_record.preAuthInitiationAmount = row[28]
                            existing_record.preAuthCancelDate = row[29]
                            existing_record.preAuthApprovalDate = row[30]
                            existing_record.preAuthApprovalAmount = row[31]
                            existing_record.preAuthRejectionDate = row[32]
                            existing_record.enhancementFlag = row[33]
                            existing_record.enhancementApprovedAmount = row[34]
                            existing_record.surgeryDate = row[35]
                            existing_record.dischargeDate = row[36]
                            existing_record.deathDate = row[37]
                            existing_record.claimRaisedDate = row[38]
                            existing_record.claimPaidAmount = row[39]
                            existing_record.cpdApprovedDate = row[40]
                            existing_record.cpdRejectedDate = row[41]
                            existing_record.shaApprovedDate = row[42]
                            existing_record.claimPaidDate = row[43]
                            existing_record.claimUTRNumber = row[44]
                            existing_record.hospitalDistrict = row[45]
                            existing_record.hospitalState = row[46]
                            existing_record.preAuthApprovalRemarks = row[47]
                            existing_record.preAuthRejectionRemarks = row[48]
                            existing_record.claimApprovalRemarks = row[49]
                            existing_record.claimRejectionRemarks = row[50]
                            existing_record.claimInitiatedAmount = row[51]
                            existing_record.rfAmount = row[52]
                            existing_record.tdsAmount = row[53]
                            existing_record.amountPaidToHospital = row[54]
                            existing_record.claimApprovedAmount = row[55]
                            existing_record.claimUpdatedDate = row[56]
                            existing_record.preauthPendingRemarks = row[57].strip()
                            existing_record.preauthPendingDate = row[58]
                            existing_record.preauthPendingUpdatedRemarks = row[59]
                            existing_record.preauthPendingUpdatedDate = row[60]
                            existing_record.claimPendingRemarks = row[61].strip()
                            existing_record.claimPendingDate = row[62]
                            existing_record.claimPendingUpdatedRemarks = row[63]
                            existing_record.claimPendingUpdatedDate = row[64]
                            existing_record.lastUpdatedUser = row[65]
                            existing_record.isAadharBenificaiary = row[66]
                            existing_record.bioAuthAtRegistration = row[67]
                            existing_record.bioAuthAtdischarge = row[68]
                            existing_record.erroneousInitiatedAmount = row[69]
                            existing_record.erroneousInitiatedDate = row[70]
                            existing_record.erroneousApprovedAmount = row[71]
                            existing_record.erroneousApprovedDate = row[72]
                            existing_record.erroneousPaidDate = row[73]
                            existing_record.erroneousUTRNumber = row[74]
                            existing_record.ipNumber = row[75]
                            existing_record.cpdPendingCount = row[76]
                            existing_record.cpdProcessingTime = row[77]
                            existing_record.revokedCase = row[78]
                            existing_record.revokedDate = row[79]
                            existing_record.revokedRemarks = row[80]
                            existing_record.insuranceLiableAmount = row[81]
                            existing_record.trustLiableAmount = row[82]
                            existing_record.patientLiableAmount = row[83]
                            existing_record.actualRegistrationDate = row[84]
                            

                            # ... update other fields as required
                            existing_record.save()
                        else:
                            
                            linked_casenumber = PreAuthLinkcaseNumber.objects.get_or_create(caseNumber = row[0] , user = request.user)
                            data_list = []
                            # Create a new record
                            data = MPClaimPaidExcel(
                                user = request.user , 
                                caseNumber_id=row[0], workflowStatus=row[1], statusUpdateDate=row[2], nhmpID=row[3], familyId=row[4],
                                patientName=row[5], age=row[6], gender=row[7], patientsHomeDistrict=row[8],
                                patientsHomeState=row[9], patientPhoneNo=row[10], address=row[11], communicationAddress=row[12],
                                communicationContactNo=row[13], communicationVillage=row[14], communicationMandal=row[15],
                                communicationDistrict=row[16], communicationState=row[17], speciality=row[18], patientIPOP=row[19], procedureCode=row[20],
                                procedure=row[21], procedureAutoApprove=row[22], medicalOrSurgery=row[23],
                                ipRegistrationDate=row[24], hospitalName=row[25], admissionDate=row[26], preAuthInitiationDate=row[27],
                                preAuthInitiationAmount=row[28],
                                preAuthCancelDate=row[29], preAuthApprovalDate=row[30], preAuthApprovalAmount=row[31], preAuthRejectionDate=row[32],
                                enhancementFlag=row[33], enhancementApprovedAmount=row[34], surgeryDate=row[35],
                                dischargeDate=row[36], deathDate=row[37], claimRaisedDate=row[38], claimPaidAmount=row[39], cpdApprovedDate=row[40],
                                cpdRejectedDate=row[41],
                                shaApprovedDate=row[42], claimPaidDate=row[43], claimUTRNumber=row[44], hospitalDistrict=row[45], hospitalState=row[46],
                                preAuthApprovalRemarks=row[47], preAuthRejectionRemarks=row[48], claimApprovalRemarks=row[49], claimRejectionRemarks=row[50],
                                claimInitiatedAmount=row[51], rfAmount=row[52], tdsAmount=row[53], amountPaidToHospital=row[54], claimApprovedAmount=row[55],
                                claimUpdatedDate=row[56], preauthPendingRemarks=row[ 57].strip(), preauthPendingDate=row[58], preauthPendingUpdatedRemarks=row[59],
                                preauthPendingUpdatedDate=row[60],
                                claimPendingRemarks=row[61].strip(), claimPendingDate=row[ 62], claimPendingUpdatedRemarks=row[63], claimPendingUpdatedDate=row[64],
                                lastUpdatedUser=row[65],
                                isAadharBenificaiary=row[66], bioAuthAtRegistration=row[67], bioAuthAtdischarge=row[68], erroneousInitiatedAmount=row[69],
                                erroneousInitiatedDate=row[70], erroneousApprovedAmount=row[71], erroneousApprovedDate=row[72], erroneousPaidDate=row[73],
                                erroneousUTRNumber=row[74], ipNumber=row[75], cpdPendingCount=row[76], cpdProcessingTime=row[77], revokedCase=row[78],
                                revokedDate=row[79], revokedRemarks=row[80], insuranceLiableAmount=row[81], trustLiableAmount=row[82],
                                patientLiableAmount=row[83], actualRegistrationDate=row[84] )
                
                            data_list.append(data)
                            MPClaimPaidExcel.objects.bulk_create(data_list)
                            

                    return Response({"status": "Success", "message": "Successfully Uploaded."})
            except Exception as e:
                # logger.critical(str(e), exc_info=True)
                error_message = str(e)
                detail_message = error_message.split(":")[-1].strip()
                if detail_message:
                    caseNumber = detail_message.split('(')[2].split(')')[0]
                return Response({"status": "error",
                                 "message": f"Error when attempting to store this Case Number : {caseNumber}, It's because this Case Number is not linked to any PreAuth."},
                                status=400)

        except Exception as e:

            return Response({"status": "error",
                             "message": " Please Check your Excel File Format and upload Again.",
                             },
                            status=400)


class DumpCSVInsert(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = DumpCSVSerializer

    def post(self, request, format=None):
        if 'csv_file' not in request.FILES:
            return Response({"status": "error", "message": "No file uploaded."}, status=400)

        csv_file = request.FILES["csv_file"]
        if csv_file.size == 0 or not csv_file.name.endswith(".csv"):
            return Response({"status": "error", "message": "Only .csv files are supported."}, status=400)

        # Convert the CSV file to a string and create a StringIO object
        csv_string = csv_file.read().decode('utf-8')
        csv_io = StringIO(csv_string)

        data_list = []
        csv_reader = csv.reader(csv_io)
        next(csv_reader)  # Skip the header row

        # Loop through the rows and create DumpExcel objects
        for row in csv_reader:
            data = MPClaimPaidExcel(
                caseNumber=row[0], workflowStatus=row[1], statusUpdateDate=row[2], nhmpID=row[3], familyId=row[4],
                patientName=row[5], age=row[6], gender=row[7], patientsHomeDistrict=row[8],
                patientsHomeState=row[9], patientPhoneNo=row[10], address=row[11], communicationAddress=row[12],
                communicationContactNo=row[13], communicationVillage=row[14], communicationMandal=row[15],
                communicationDistrict=row[16], communicationState=row[
                    17], speciality=row[18], patientIPOP=row[19], procedureCode=row[20],
                procedure=row[21], procedureAutoApprove=row[22], medicalOrSurgery=row[23],
                ipRegistrationDate=row[24], hospitalName=row[25], admissionDate=row[
                    26], preAuthInitiationDate=row[27], preAuthInitiationAmount=row[28],
                preAuthCancelDate=row[29], preAuthApprovalDate=row[
                    30], preAuthApprovalAmount=row[31], preAuthRejectionDate=row[32],
                enhancementFlag=row[33], enhancementApprovedAmount=row[34], surgeryDate=row[35],
                dischargeDate=row[36], deathDate=row[37], claimRaisedDate=row[
                    38], claimPaidAmount=row[39], cpdApprovedDate=row[40], cpdRejectedDate=row[41],
                shaApprovedDate=row[42], claimPaidDate=row[43], claimUTRNumber=row[
                    44], hospitalDistrict=row[45], hospitalState=row[46],
                preAuthApprovalRemarks=row[47], preAuthRejectionRemarks=row[
                    48], claimApprovalRemarks=row[49], claimRejectionRemarks=row[50],
                claimInitiatedAmount=row[51], rfAmount=row[52], tdsAmount=row[
                    53], amountPaidToHospital=row[54], claimApprovedAmount=row[55],
                claimUpdatedDate=row[56], preauthPendingRemarks=row[57], preauthPendingDate=row[
                    58], preauthPendingUpdatedRemarks=row[59], preauthPendingUpdatedDate=row[60],
                claimPendingRemarks=row[61], claimPendingDate=row[62], claimPendingUpdatedRemarks=row[
                    63], claimPendingUpdatedDate=row[64], lastUpdatedUser=row[65],
                isAadharBenificaiary=row[66], bioAuthAtRegistration=row[
                    67], bioAuthAtdischarge=row[68], erroneousInitiatedAmount=row[69],
                erroneousInitiatedDate=row[70], erroneousApprovedAmount=row[
                    71], erroneousApprovedDate=row[72], erroneousPaidDate=row[73],
                erroneousUTRNumber=row[74], ipNumber=row[75], cpdPendingCount=row[
                    76], cpdProcessingTime=row[77], revokedCase=row[78],
                revokedDate=row[79], revokedRemarks=row[80], insuranceLiableAmount=row[81], trustLiableAmount=row[82],
                patientLiableAmount=row[83], actualRegistrationDate=row[84],)

            data_list.append(data)

        # Use bulk_create to insert all the data into the database
        MPClaimPaidExcel.objects.bulk_create(data_list)

        return Response({"status": "Success", "message": "Successfully Uploaded."})


class DumpExcelView(generics.ListAPIView):
    pagination_class = LimitOffsetPagination
    serializer_class = DumpExcelViewSerializer

    def get_queryset(self):
        try:
            queryset = DumpExcel.objects.all().order_by('-actualRegistrationDate')
            # queryset = DumpExcel.objects.filter(actualRegistrationDate__range= ['2020-07-01' , '2020-08-05'])
        except DumpExcel.DoesNotExist:
            return Response(status=400)
        return queryset

    def get(self, request, start_date=None, end_date=None):
        # queryset = DumpExcel.objects.filter(actualRegistrationDate__range= '2022')
        page = self.paginate_queryset(self.get_queryset())
        if page is not None:
            serializer = self.get_serializer(page, many=True).data
            return self.get_paginated_response({'status': 'success',
                                                'message': 'Data fetched successfully',
                                                'data': serializer})

        serializer = self.get_serializer(self.get_queryset(), many=True).data
        return Response({'status': 'success',
                        'message': 'Data fetched successfully',
                         'data': serializer})


def get_financial_year(date):
            if date.month >= 4:
                return f"{date.year}-{date.year+1}"
            else:
                return f"{date.year-1}-{date.year}"

class FinalExcel(APIView):
    def get(self, request):
        data_list = [['PreAuth','CaseNumber', 'NHMPID', 'patientName', 'PatientDistrict', 'Gender', 'Age', 'categoryDetails', 'procedureDetails',
                     'caseType', 'isAadharBenificaiary', 'bioAuthAtRegistration', 'bioAuthAtdischarge', 'caseStatus', 'hospitalName',
                      'hospitalDistrict', 'admissionDate', 'preauthDate', 'preauthApproveDate', 'Month - Preauth ', 'Financial Year - Preauth', 'Financial Year - Preauth',
                    'preauthApprovedAmount','preauthRejectedDate', 'surgeryDate', 'deathDate',  'dischargeDate', 'Days Since Discharge',
                      'Preauth Pending Remarks', 'preauthPendingUpdatedDate', 'claim Pending Remarks', 'claimPendingUpdatedDate', 'Claim Rejection Remarks', 'claimSubmittedDate', 'claimInitaiatedAmount',
                      'claimApprovedAmount', 'Claim Paid Date', 'Month - Claim Paid' ,'Calender Year - Claim Paid ', 'Calender Year - Claim Paid' , 'Erroneous Initiated Date',
                      'Erroneous Initiated Amount', 'Erroneous Approved Amount', 'Erroneous Paid Date']]

        preauth_link = PreAuthLinkcaseNumber.objects.select_related('preAuthID').all()
        for i in preauth_link:
            
            dumpexcel_fields = DumpExcel.objects.filter(caseNo__caseNumber=i.caseNumber)
            MPClaimPaidExcel_fields = MPClaimPaidExcel.objects.filter(caseNumber__caseNumber=i.caseNumber)
        
            for MProw, row in zip(MPClaimPaidExcel_fields, dumpexcel_fields):
                try:
                    days_since_discharge = str((datetime.datetime.today() - row.dischargeDate).days) + ' days'
                except:
                    days_since_discharge = ''

                preauth_date_month = calendar.month_name[row.preauthDate.month] if isinstance(row.preauthDate, datetime.datetime) else ''
                preauth_date_year = row.preauthDate.year if isinstance(row.preauthDate, datetime.datetime) else ''
                preauth_financial_year = get_financial_year(row.preauthDate) if isinstance(row.preauthDate, datetime.datetime) else ''

                claim_paid_date_month = calendar.month_name[MProw.claimPaidDate.month] if isinstance(MProw.claimPaidDate, datetime.datetime) else ''
                claim_paid_date_year = MProw.claimPaidDate.year if isinstance(MProw.claimPaidDate, datetime.datetime) else ''
                claim_paid_financial_year = get_financial_year(MProw.claimPaidDate) if isinstance(MProw.claimPaidDate, datetime.datetime) else ''
                if i.preAuthID == None:
                    PreAuth_ID = ''
                else:
                    PreAuth_ID = i.preAuthID.preAuthID
                data_list.append([
                    PreAuth_ID, i.caseNumber, MProw.nhmpID, MProw.patientName, row.patientDistrict, MProw.gender, row.age, 
                    row.procedureDetails, row.caseType, MProw.isAadharBenificaiary, MProw.bioAuthAtRegistration, MProw.bioAuthAtdischarge, row.caseStatus,
                    row.hospitalName, row.hospitalDistrict,
                    row.admissionDate.strftime('%d/%m/%Y %I:%M:%S %p') if row.admissionDate else '',
                    row.preauthDate.strftime('%d/%m/%Y %I:%M:%S %p') if row.preauthDate else '',
                    row.preauthApproveDate.strftime('%d/%m/%Y %I:%M:%S %p') if row.preauthApproveDate else '',
                    preauth_date_month,  
                    preauth_date_year,
                    preauth_financial_year,
                    row.preauthApprovedAmount,
                    row.preauthRejectedDate.strftime('%d/%m/%Y %I:%M:%S %p') if row.preauthRejectedDate else '',
                    row.surgeryDate.strftime('%d/%m/%Y %I:%M:%S %p') if row.surgeryDate else '',
                    row.deathDate.strftime('%d/%m/%Y %I:%M:%S %p') if row.deathDate else '',
                    row.dischargeDate.strftime('%d/%m/%Y %I:%M:%S %p') if row.dischargeDate else '',
                    days_since_discharge,
                    MProw.preauthPendingRemarks,
                    MProw.preauthPendingUpdatedDate.strftime('%d/%m/%Y %I:%M:%S %p') if MProw.preauthPendingUpdatedDate else '',
                    MProw.claimPendingRemarks,
                    MProw.claimPendingUpdatedDate.strftime('%d/%m/%Y %I:%M:%S %p') if MProw.claimPendingUpdatedDate else '',
                    MProw.claimRejectionRemarks,
                    row.claimSubmittedDate.strftime('%d/%m/%Y %I:%M:%S %p') if row.claimSubmittedDate else '',
                    row.claimInitaiatedAmount,
                    row.claimApprovedAmount,
                    MProw.claimPaidDate.strftime('%d/%m/%Y %I:%M:%S %p') if MProw.claimPaidDate else '',
                    claim_paid_date_month,
                    claim_paid_date_year,
                    claim_paid_financial_year,
                    MProw.erroneousInitiatedDate.strftime('%d/%m/%Y %I:%M:%S %p') if MProw.erroneousInitiatedDate else '',
                    MProw.erroneousInitiatedAmount,
                    MProw.erroneousApprovedAmount,
                    MProw.erroneousPaidDate.strftime('%d/%m/%Y %I:%M:%S %p') if MProw.erroneousPaidDate else '',
                ])

        wb = openpyxl.Workbook()
        ws = wb.active
        for row in data_list:   
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="data.xlsx"'
        wb.save(response)
        return response

